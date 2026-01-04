
import openpyxl
from datetime import datetime

def create_and_parse():
    # Create the workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Headers
    headers = ["fecha", "nombre documento", "num doc", "tercero", "nit", "cod cuenta", "nombre cuenta", "concepto", "debito", "credito"]
    ws.append(headers)
    
    # Data Row
    # Note: Using strings for numbers as they often appear in imports, checking robustness
    row_data = [
        datetime(2026, 1, 11), 
        "Comprobante de Egreso", 
        "777999", 
        "Jaime Muñoz", 
        "41411660", 
        "513506", 
        "Servicio de energía", 
        "energia mes", 
        777888, 
        0.00
    ]
    ws.append(row_data)
    
    filename = "debug_reproduce.xlsx"
    wb.save(filename)
    print(f"Created {filename}")
    
    # Parse it back
    print("--- Parsing ---")
    wb_read = openpyxl.load_workbook(filename, data_only=True)
    sheet = wb_read.active
    
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    for i, row in enumerate(rows):
        print(f"Row {i} raw: {row}")
        print(f"Index 0 (A): {row[0]}")
        print(f"Index 2 (C): {row[2]}")
        
if __name__ == "__main__":
    create_and_parse()
