import openpyxl
from datetime import date, datetime
from typing import List, Dict, Any, Optional
from sqlalchemy.orm import Session
from app.models import Documento, MovimientoContable, PlanCuenta, Tercero, TipoDocumento
from app.services.import_utils import ImportUtils

class UniversalImportService:

    @staticmethod
    def get_headers(file_content: bytes) -> List[str]:
        """
        Extrae la primera fila (encabezados) de un archivo Excel o CSV.
        Utilitario para la configuración de plantillas.
        """
        from io import BytesIO, StringIO
        import csv
        import openpyxl

        # Intentar Excel
        try:
            workbook = openpyxl.load_workbook(filename=BytesIO(file_content), data_only=True)
            sheet = workbook.active
            if sheet.max_row >= 1:
                return [str(cell.value or "").strip() for cell in sheet[1]]
            return []
        except Exception:
            # Fallback CSV
            try:
                text = file_content.decode('utf-8-sig')
            except:
                text = file_content.decode('latin-1', errors='ignore')

            # Sniffer
            try:
                sample = text[:2048]
                sniffer = csv.Sniffer()
                dialect = sniffer.sniff(sample, delimiters=';,\t|')
                sep = dialect.delimiter
            except:
                sep = ',' if text.count(',') > text.count(';') else ';'
            
            f = StringIO(text)
            reader = csv.reader(f, delimiter=sep)
            rows = list(reader)
            if rows:
                return [str(c).strip() for c in rows[0]]
            return []

    @staticmethod
    def parse_excel_file(file_content: bytes, mapping_config: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        """
        Lee el archivo (Excel o CSV) y retorna una lista de diccionarios planos.
        Soporta .xlsx (openpyxl) y .csv (utf-8/latin-1).
        Si se provee mapping_config, usa esa configuración para extraer columnas.
        """
        from io import BytesIO, StringIO
        import csv
        import codecs

        entries = []
        is_csv = False
        rows_iter = []

        # Intentar cargar como Excel
        try:
            workbook = openpyxl.load_workbook(filename=BytesIO(file_content), data_only=True)
            sheet = workbook.active
            # Read headers (first row)
            header_row_excel = [str(cell.value).lower() for cell in sheet[1]] if sheet.max_row >= 1 else []
            rows_iter = list(sheet.iter_rows(min_row=2, values_only=True))
        except Exception:
            # Fallback a CSV
            is_csv = True
            try:
                # Detectar encoding (utf-8-sig para BOM)
                text = file_content.decode('utf-8-sig')
            except:
                text = file_content.decode('latin-1', errors='ignore')
            
            # Detectar separador usando Sniffer para mayor robustez
            try:
                # Tomar una muestra de las primeras lineas
                sample = text[:2048]
                sniffer = csv.Sniffer()
                # Permitir delimitadores comunes
                dialect = sniffer.sniff(sample, delimiters=';,\t|')
                sep = dialect.delimiter
            except:
                # Fallback manual si Sniffer falla
                sep = ',' if text.count(',') > text.count(';') else ';'
            
            f = StringIO(text)
            reader = csv.reader(f, delimiter=sep)
            
            # Saltar header
            all_rows = list(reader)
            if len(all_rows) > 0:
                header_row_csv = [str(cell).lower() for cell in all_rows[0]]
                rows_iter = all_rows[1:] # Asumimos fila 1 es header
            else:
                rows_iter = []

        
        # Iterar filas normalizadas
        # Detectar formato basado en headers (fila 0 de rows_iter, o adivinar por contenido)
        # Formato Standard (9 cols): Fecha, Tipo, Num, Cta, Nit, NomTer, Det, Deb, Cred
        # Formato User (8 cols): Fecha, Documento, Beneficiario, CodCta, NomCta, Concepto, Deb, Cred
        
        is_user_format = False
        is_extended_format = False 
        is_journal_format = False
        is_smart_export_format = False
        
        if not mapping_config:
            # AUTO-DETECTION LOGIC (only if no template provided)
            header_row = None
            if 'header_row_excel' in locals() and header_row_excel:
                header_row = header_row_excel
            elif 'header_row_csv' in locals() and header_row_csv:
                header_row = header_row_csv
            elif rows_iter and 'all_rows' in locals() and len(all_rows) > 0:
                header_row = [str(cell).lower() for cell in all_rows[0]]
            
            if header_row:
                 print(f"[DEBUG IMPORT] Headers: {header_row}")
                 
                 # SMART EXPORT DETECTION (Relaxed)
                 # Check for "Nombre Tipo" OR just "Tipo" and "Nombre" 
                 if any('nombre tipo' in h for h in header_row) or (any('tipo' in h for h in header_row) and any('nombre' in h for h in header_row) and len(header_row) >= 11):
                     is_smart_export_format = True
                     print(f"[DEBUG IMPORT] Detected SMART EXPORT format. Headers: {header_row}")
                 
                 if any('documento' in h for h in header_row) and not any('tipo' in h for h in header_row) and not is_smart_export_format:
                     is_user_format = True
                 elif len(header_row) == 8 and not is_smart_export_format:
                     is_user_format = True
                 
                 # Extended detection via Headers
                 elif (any('nombre cuenta' in h for h in header_row) or any('nom cta' in h for h in header_row)) and not is_smart_export_format:
                     if len(header_row) == 10:
                         is_journal_format = True
                         print("[DEBUG IMPORT] Detected JOURNAL format")
                     else:
                         is_extended_format = True
                         print("[DEBUG IMPORT] Detected EXTENDED format")

            elif rows_iter and len(rows_iter[0]) == 8:
                 # Fallback: Count cols of first data row
                 is_user_format = True
                 print("[DEBUG IMPORT] Fallback: USER format detected by col count (8)")
                 
            # Fallback length check if headers didn't decide
            if not is_extended_format and not is_user_format and not is_journal_format and not is_smart_export_format and rows_iter:
                 row_len = len(rows_iter[0])
                 print(f"[DEBUG IMPORT] Fallback check. Row Len: {row_len}")
                 if row_len >= 11:
                      # Ambiguity: 11 cols could be Extended OR Smart. 
                      # Smart usually has 'nombre tipo' at col 2. Extended has Num at col 2.
                      # We check content of col 2 of first row. If it looks like a number, likely Extended. If text, Smart.
                      sample_val = str(rows_iter[0][2])
                      if not any(char.isdigit() for char in sample_val) and len(sample_val) > 4:
                          is_smart_export_format = True
                          print("[DEBUG IMPORT] Fallback: SMART EXPORT detected by 11+ cols and text in col 2")
                      else:
                          is_extended_format = True
                          print("[DEBUG IMPORT] Fallback: EXTENDED format detected by 11+ cols")
                 elif row_len == 10:
                      is_journal_format = True
                      print("[DEBUG IMPORT] Fallback: JOURNAL format detected by 10 cols")

        for row_idx, row in enumerate(rows_iter, start=2):
            if not row: continue
            
            # Normalizar longitud row para evitar index error (Expand to 50 cols just in case)
            if len(row) < 50:
                row = list(row) + [None]*(50-len(row))
            
            # --- HEURISTICS: Runtime Format Switch ---
            # If we detected EXTENDED but row[2] is clearly text (e.g. "COMPROBANTE...") and row[3] is number
            # Then we made a mistake and it IS Smart Export.
            if is_extended_format and not is_smart_export_format:
                val_2 = str(row[2]).strip()
                val_3 = str(row[3]).strip()
                # Check if val_2 is NOT numeric (has letters) and val_3 IS numeric
                if (any(c.isalpha() for c in val_2)) and (any(c.isdigit() for c in val_3)):
                    print(f"[DEBUG IMPORT] SWITCHING to SMART EXPORT based on row content (2='{val_2}', 3='{val_3}')")
                    is_extended_format = False
                    is_smart_export_format = True
            # ------------------------------------------

            # Limpieza básica
            check_col = mapping_config.get('fecha', 0) if mapping_config else 0
            if not row[check_col]: continue 
            
            try:
                # Fecha
                raw_fecha = row[check_col]
                fecha_obj = None

                if isinstance(raw_fecha, datetime):
                    fecha_obj = raw_fecha.date()
                elif isinstance(raw_fecha, str):
                    # Limpiar comillas extras
                    raw_fecha = raw_fecha.replace('"', '').replace("'", "").strip()
                    # Formatos comunes
                    fmt_list = ['%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y']
                    for fmt in fmt_list:
                        try:
                            fecha_obj = datetime.strptime(raw_fecha, fmt).date()
                            break
                        except:
                            pass
                
                if not fecha_obj: continue

                def clean_str(val):
                    return str(val).strip() if val is not None else ""

                def clean_float(val):
                    if isinstance(val, (int, float)): return float(val)
                    if isinstance(val, str):
                        v = val.replace('$', '').strip()
                        if not v: return 0.0
                        if ',' in v and '.' in v:
                             if v.rfind(',') > v.rfind('.'): v = v.replace('.', '').replace(',', '.')
                             else: v = v.replace(',', '')
                        elif ',' in v: v = v.replace(',', '.')
                        try: return float(v)
                        except: return 0.0
                    return 0.0

                if mapping_config:
                    # --- USE TEMPLATE MAPPING ---
                    # Helper to get value safely
                    def get_val(key, default=None):
                        idx = mapping_config.get(key)
                        if idx is None: return default
                        try: 
                            val = row[int(idx)]
                            return val
                        except: return default
                    
                    # Extraer campos directos
                    tipo_code = clean_str(get_val('tipo_doc'))
                    tipo_name = clean_str(get_val('nombre_tipo_doc'))
                    numero_val = clean_str(get_val('numero'))
                    
                    if not tipo_code and tipo_name:
                        tipo_code = tipo_name
                    if not tipo_code and not tipo_name:
                         tipo_code = "GEN"

                    entry = {
                        "fecha": fecha_obj,
                        "tipo_doc": tipo_code,
                        "nombre_tipo_doc": tipo_name,
                        "numero": numero_val or "1",
                        "cuenta": clean_str(get_val('cuenta')),
                        "nombre_cuenta": clean_str(get_val('nombre_cuenta')),
                        "nit": clean_str(get_val('nit')),
                        "nombre_tercero": clean_str(get_val('tercero') or get_val('nombre_tercero')).upper(),
                        "detalle": clean_str(get_val('detalle')) or "Importación Plantilla",
                        "debito": clean_float(get_val('debito')),
                        "credito": clean_float(get_val('credito'))
                    }

                elif is_smart_export_format:
                     # Mapping SMART EXPORT (Corrected Mappings for Shifted Cols)
                     # 0=Fecha, 1=Tipo, 2=NomTipo, 3=Num, 4=NameTer, 5=Nit, 6=Cta, 7=NomCta, 8=Detalle, 9=Deb, 10=Cred
                     entry = {
                        "fecha": fecha_obj,
                        "tipo_doc": clean_str(row[1]) or "GEN",
                        "nombre_tipo_doc": clean_str(row[2]),
                        "numero": clean_str(row[3]),
                        "cuenta": clean_str(row[6]),  # Shifted to 6
                        "nombre_cuenta": clean_str(row[7]), # Shifted to 7
                        "nit": clean_str(row[5]),
                        "nombre_tercero": clean_str(row[4]).upper(), # Name at 4
                        "detalle": clean_str(row[8]) or "Importación Automática",
                        "debito": clean_float(row[9]), # Might need check if these shifted too
                        "credito": clean_float(row[10]) # Assuming Deb/Cred are at end
                     }

                elif is_user_format:
                    # Mapping Formato Usuario (8 cols)
                    raw_doc = clean_str(row[1])
                    import re
                    match = re.match(r"([A-Za-z\s]+).*?(\d+)", raw_doc)
                    if match:
                        tipo_val = match.group(1).replace("#", "").strip()
                        num_val = match.group(2)
                    else:
                        tipo_val = raw_doc[:3]
                        num_val = "".join(filter(str.isdigit, raw_doc))
                    
                    benef_name = clean_str(row[2])
                    
                    entry = {
                        "fecha": fecha_obj,
                        "tipo_doc": tipo_val or "GEN",
                        "numero": num_val,
                        "cuenta": clean_str(row[3]),
                        "nit": "", 
                        "nombre_tercero": benef_name.upper(),
                        "detalle": clean_str(row[5]) or "Importación",
                        "debito": clean_float(row[6]),
                        "credito": clean_float(row[7])
                    }
                elif is_extended_format:
                    # Mapping Extended (11 cols)
                    entry = {
                        "fecha": fecha_obj,
                        "tipo_doc": clean_str(row[1]) or "GEN",
                        "numero": clean_str(row[2]),
                        "cuenta": clean_str(row[5]),
                        "nit": clean_str(row[4]),
                        "nombre_tercero": clean_str(row[3]).upper(),
                        "detalle": clean_str(row[7]) or "Importación Universal",
                        "debito": clean_float(row[9]),
                        "credito": clean_float(row[10]) 
                    }
                elif is_journal_format:
                    entry = {
                        "fecha": fecha_obj,
                        "tipo_doc": clean_str(row[1]) or "GEN",
                        "numero": clean_str(row[2]),
                        "cuenta": clean_str(row[5]),
                        "nit": clean_str(row[4]),
                        "nombre_tercero": clean_str(row[3]).upper(),
                        "detalle": clean_str(row[7]) or "Importación Diario",
                        "debito": clean_float(row[8]),
                        "credito": clean_float(row[9])
                    }
                else:
                    # Mapping Standard (9 cols)
                    # 0=Fecha, 1=Tipo, 2=Num, 3=Cta, 4=Nit, 5=NomTer, 6=Det, 7=Deb, 8=Cred
                    entry = {
                        "fecha": fecha_obj,
                        "tipo_doc": clean_str(row[1]) or "GEN",
                        "numero": clean_str(row[2]),
                        "cuenta": clean_str(row[3]),
                        "nit": clean_str(row[4]),
                        "nombre_tercero": clean_str(row[5]).upper() or "TERCERO GENERICO",
                        "detalle": clean_str(row[6]) or "Importación Universal",
                        "debito": clean_float(row[7]),
                        "credito": clean_float(row[8])
                    }
                
                # Validar cuenta obligatoria
                if not entry["cuenta"]: continue
                
                entries.append(entry)
                
            except Exception as e:
                print(f"Error parsing row {row_idx}: {e}")
                continue
                
        return entries

    @staticmethod
    def process_import(db: Session, empresa_id: int, file_content: bytes, default_tercero_id: Optional[int] = None, template_id: Optional[int] = None) -> Dict[str, Any]:
        """
        Procesa el archivo Excel y crea la contabilidad.
        """
        mapping_config = None
        if template_id:
            print(f"[DEBUG] Using Template ID: {template_id}")
            from app.models import ImportTemplate
            tpl = db.query(ImportTemplate).filter_by(id=template_id, empresa_id=empresa_id).first()
            if tpl:
                mapping_config = tpl.mapping_config

        entries = UniversalImportService.parse_excel_file(file_content, mapping_config)
        results = {
            "accounts": 0, 
            "third_parties": 0, 
            "documents": 0, 
            "transactions": 0, 
            "errors": [],
            "created_documents": [] # New: Detailed summary for UI
        }
        
        # 1. Verificar Límites de Plan + Cupos Adicionales
        from app.models import Empresa
        
        # Refrescar instancia para asegurar que tenemos el dato más reciente (por si el usuario acaba de ampliar)
        empresa = db.query(Empresa).filter_by(id=empresa_id).first()
        db.refresh(empresa) 

        base_limit = empresa.limite_registros
        
        if base_limit is not None:
             # Calcular cupo extra (Sumamos todos los cupos adicionales históricos/activos)
             # Asumimos que si compran cupo, es "capacidad añadida" al total.
             extra_quota = sum(c.cantidad_adicional for c in empresa.cupos_adicionales)
             total_limit = base_limit + extra_quota
             
             current_usage = db.query(MovimientoContable).join(Documento).filter(Documento.empresa_id == empresa_id).count()
             remaining_quota = total_limit - current_usage
             
             if remaining_quota <= 0:
                 results["errors"].append(f"Límite de registros excedido ({current_usage}/{total_limit}). Actualice su plan o adquiera cupos adicionales.")
                 return results
        else:
             remaining_quota = 999999999

        # 1. Agrupar por Documento (llave: Tipo + Numero + Fecha)
        grouped_docs = {}
        
        # caches
        acc_cache = {} # code -> id
        tercero_cache = {} # nit -> id
        tipo_doc_cache = {} # code -> id
        
        # caches
        acc_cache = {} # code -> id
        tercero_cache = {} # nit -> id
        tipo_doc_cache = {} # code -> id
        
        # --- PHASE 1: PRE-SCAN & RESOLVE DOCUMENT TYPES ---
        # Objetivo: Resolver todos los Nombres de Documento a Códigos de DB *antes* de agrupar.
        unique_type_names = set()
        for e in entries:
            if e.get('nombre_tipo_doc'):
                unique_type_names.add(e['nombre_tipo_doc'])
                
        name_to_code_map = {} # "Recibo de Caja" -> "RC"
        
        for t_name in unique_type_names:
            # Asegurar existencia (crea si falta)
            td_id = ImportUtils.ensure_tipo_documento_exists(db, empresa_id, nombre=t_name)
            if td_id:
                td_obj = db.query(TipoDocumento).filter_by(id=td_id).first()
                if td_obj:
                    name_to_code_map[t_name] = td_obj.codigo
                    tipo_doc_cache[td_obj.codigo] = td_id # Pre-calentar cache

        # --- PHASE 2: GROUPING ---
        # Agrupar usando el CÓDIGO REAL de la base de datos
        grouped_docs = {}
        
        for e in entries:
            t_name = e.get('nombre_tipo_doc')
            t_code = e.get('tipo_doc')
            
            # Si tenemos un mapeo por nombre, es la autoridad.
            # Esto separa "Recibo #1" de "Comprobante #1" aunque venían sin código.
            resolved_code = t_code
            if t_name and t_name in name_to_code_map:
                resolved_code = name_to_code_map[t_name]
            
            # Ajuste en el entry para el procesamiento posterior
            e['tipo_doc'] = resolved_code 
            
            key = (resolved_code, e['numero'], e['fecha'])
            if key not in grouped_docs: grouped_docs[key] = []
            grouped_docs[key].append(e)
              
        # 3. Procesar Documentos
        for (tipo_code, num_str, doc_date), lines in grouped_docs.items():
            try:
                # Z. Verificar Cupo Disponible
                doc_cost = len(lines)
                if remaining_quota < doc_cost:
                    results["errors"].append(f"Doc {tipo_code}-{num_str} saltado: Supera el límite de registros disponible ({remaining_quota}).")
                    continue 

                # A. Obtener Tipo Doc (Ya resuelto en Fase 1)
                tipo_name_sugerido = lines[0].get('nombre_tipo_doc') # Always define this
                tipo_id = tipo_doc_cache.get(tipo_code)
                
                # Fallback de seguridad (por si acaso llegó un código huerfano no escaneado)
                if not tipo_id:
                     tipo_id = ImportUtils.ensure_tipo_documento_exists(
                        db, empresa_id, codigo=tipo_code, nombre=tipo_name_sugerido
                     )
                     if tipo_id: tipo_doc_cache[tipo_code] = tipo_id

                if not tipo_id:
                     results["errors"].append(f"Doc {tipo_code}-{num_str} saltado: No se pudo determinar Tipo Documento.")
                     continue
                
                # B. Validar/Crear Tercero del Encabezado (Tomamos el de la primera línea o el mayoritario)
                # Ojo: Cada línea puede tener un tercero. El documento tiene un Beneficiario principal.
                # Usaremos el de la primera línea como Beneficiario del Documento.
                first_line = lines[0]
                header_nit = first_line['nit']
                header_name = first_line['nombre_tercero']
                
                if header_nit not in tercero_cache:
                    t_id, t_created = ImportUtils.ensure_tercero_exists(db, empresa_id, header_nit, header_name, default_tercero_id)
                    tercero_cache[header_nit] = t_id
                    if t_created: results["third_parties"] += 1
                
                beneficiario_id = tercero_cache[header_nit]
                
                if not beneficiario_id:
                     results["errors"].append(f"Doc {tipo_code}-{num_str} saltado: Sin beneficiario válido.")
                     continue

                # C. Parsear Numero
                # Validation: If num_str is not numeric or contains no digits, warn user instead of 999999
                try: 
                    digits_only = ''.join(filter(str.isdigit, str(num_str)))
                    if not digits_only:
                         raise ValueError("No digits found")
                    doc_num = int(digits_only)
                except Exception: 
                    # Fix: Use 'lines' because 'row' is not defined in this scope. 
                    # Dump the first line of the group as unique sample
                    try:
                        sample_line = lines[0] if lines else {}
                        row_preview = str(sample_line)
                    except: 
                        row_preview = "Error extracting preview"
                        
                    results["errors"].append(f"Fila (Grupo): La columna 'Número' tiene un valor inválido ('{num_str}'). Verifique que la columna 'Número' contenga dígitos. Muestra: [{row_preview}]")
                    continue
                    
                if doc_num == 0: 
                     # Handle explicit 0 if needed, but usually 0 is not a valid doc number? 
                     # Let's allow 0 if it exists, or warn. 
                     # But previously fallback was 999999.
                     # Let's assume 0 is invalid for now.
                     results["errors"].append(f"Fila saltada: El número de documento no puede ser 0 ('{num_str}').")
                     continue

                # C2. Verificar Duplicidad (Idempotencia)
                # SOLO verificamos si existe un documento ACTIVO (no anulado).
                # Esto permite re-importar (reutilizar número) si el anterior fue anulado/eliminado soft.
                exists = db.query(Documento).filter_by(
                    empresa_id=empresa_id,
                    tipo_documento_id=tipo_id,
                    numero=doc_num,
                    anulado=False
                ).first()
                if exists:
                    results["errors"].append(f"Doc {tipo_code}-{doc_num} saltado: Ya existe (Activo).")
                    continue

                # D. Crear Documento
                new_doc = Documento(
                    empresa_id=empresa_id,
                    beneficiario_id=beneficiario_id,
                    tipo_documento_id=tipo_id,
                    numero=doc_num,
                    fecha=doc_date,
                    observaciones=f"Import Universal: {first_line['detalle']}",
                    estado="ACTIVO"
                )
                db.add(new_doc)
                db.flush()
                results["documents"] += 1
                
                # E. Procesar Movimientos
                for line in lines:
                    # 1. Cuenta
                    c_code = line['cuenta']
                    if c_code not in acc_cache:
                        # Crear jerarquía si no existe
                        # Nota: ImportUtils.ensure_account_hierarchy crea padre si falta.
                        # Usamos el nombre provisto en el Excel o un fallback
                        c_name = line.get('nombre_cuenta') or f"CTA IMP {c_code}"
                        
                        a_id, created_count = ImportUtils.ensure_account_hierarchy(db, empresa_id, c_code, c_name)
                        acc_cache[c_code] = a_id
                        results["accounts"] += created_count
                    
                    acc_id = acc_cache[c_code]
                    if not acc_id: continue

                    # 2. Tercero (NIT) de la línea (o del encabezado)
                    row_nit = line.get('nit') or header_nit
                    row_tercero_name = line.get('nombre_tercero') or header_name
                    
                    # Resolver ID del tercero para el movimiento
                    mov_tercero_id = beneficiario_id # Default header
                    if row_nit and row_nit != header_nit:
                         if row_nit not in tercero_cache:
                              tid, _ = ImportUtils.ensure_tercero_exists(db, empresa_id, row_nit, row_tercero_name)
                              tercero_cache[row_nit] = tid
                         mov_tercero_id = tercero_cache[row_nit]

                    # 3. Crear Movimiento
                    mov = MovimientoContable(
                        documento_id=new_doc.id,
                        cuenta_id=acc_id,
                        tercero_id=mov_tercero_id, # ASIGNANDO EL TERCERO AL MOVIMIENTO
                        concepto=line['detalle'],
                        debito=line['debito'],
                        credito=line['credito'],
                    )
                    db.add(mov)
                    results["transactions"] += 1
                
                db.commit()
                remaining_quota -= doc_cost
                
                # --- NEW: Append to Created Documents Summary ---
                # We will now collect detailed movements for the report
                for line in lines:
                    c_code = line.get('cuenta')
                    c_name = line.get('nombre_cuenta') or "" # We might need to fetch the actual name if not in line
                    
                    # Logic to get the row-specific third party or fallback to header
                    row_nit = line.get('nit') or header_nit
                    row_tercero = line.get('nombre_tercero') or header_name
                    
                    results["created_documents"].append({
                        "fecha": doc_date.strftime("%Y-%m-%d"),
                        "tipo_doc": tipo_code,
                        "nombre_tipo_doc": tipo_name_sugerido or tipo_code, # Use the suggested name or code
                        "numero": doc_num,
                        "cuenta_codigo": c_code,
                        "cuenta_nombre": c_name,
                        "nit": row_nit,
                        "tercero_nombre": row_tercero,
                        "debito": line['debito'],
                        "credito": line['credito'],
                        "detalle": line['detalle']
                    })
                # ------------------------------------------------
                # ------------------------------------------------
                
            except Exception as e:
                db.rollback()
                # Limpiar cachés para evitar usar IDs muertos (eliminados por el rollback)
                acc_cache.clear()
                tercero_cache.clear()
                tipo_doc_cache.clear()
                results["errors"].append(f"Error en Doc {tipo_code}-{num_str}: {str(e)}")
        
        return results
